"""
excel_generator.py — Professional Excel Output

Generates two types of output:
1. Invoice Summary Workbook — human-readable multi-sheet workbook
   - Summary dashboard (all invoices)
   - Individual invoice detail tabs
   - Extraction log

2. Journal Entry Upload — matches Accuron's specific ERP upload format
   (56-column Journal Entry format from their template)
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional
import csv

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers,
)
from openpyxl.utils import get_column_letter

from src.models import Invoice, ValidationStatus, ExtractedField

def _load_vendor_master() -> list[dict]:
    """Load the vendor configuration from CSV."""
    config_path = Path("config/vendor_master.csv")
    if not config_path.exists():
        return []
    vendors = []
    with open(config_path, mode='r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            vendors.append(row)
    return vendors

logger = logging.getLogger(__name__)

# ─── Styling Constants ────────────────────────────────────────────────────────

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CURRENCY_FORMAT = '#,##0.00'
BOLD_FONT = Font(name='Calibri', bold=True, size=11)
NORMAL_FONT = Font(name='Calibri', size=11)

# ERP specific theme styling
ERP_HEADER_FILL = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
ERP_HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)


def _style_erp_header_row(ws, row_num, col_count):
    """Apply Accuron ERP template header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = ERP_HEADER_FONT
        cell.fill = ERP_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



def _auto_width(ws, min_width=8, max_width=50):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, min(cell_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _style_header_row(ws, row_num, col_count):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _status_fill(status: ValidationStatus):
    """Get fill color for validation status."""
    return {
        ValidationStatus.PASS: PASS_FILL,
        ValidationStatus.WARNING: WARN_FILL,
        ValidationStatus.FAIL: FAIL_FILL,
    }.get(status, WARN_FILL)


# ─── Summary Sheet ────────────────────────────────────────────────────────────

def _write_summary_sheet(ws, invoices: list[Invoice]):
    """Write summary dashboard with all invoices."""
    ws.title = "Summary"

    headers = [
        "S.No", "Source File", "PDF Type", "Invoice Number", "Invoice Date",
        "Seller Name", "Seller GSTIN", "Buyer GSTIN",
        "Subtotal (₹)", "Total Tax (₹)", "Grand Total (₹)",
        "Line Items", "Validation Status", "Confidence", "Errors",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = 'A2'

    for i, inv in enumerate(invoices, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=inv.source_file)
        ws.cell(row=row, column=3, value=inv.pdf_type.value)
        ws.cell(row=row, column=4, value=str(inv.invoice_number))
        ws.cell(row=row, column=5, value=str(inv.invoice_date))
        ws.cell(row=row, column=6, value=inv.seller.name)
        ws.cell(row=row, column=7, value=inv.seller.gstin)
        ws.cell(row=row, column=8, value=inv.buyer.gstin)
        ws.cell(row=row, column=9, value=inv.subtotal).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=10, value=inv.total_tax).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=11, value=inv.grand_total_float).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=12, value=len(inv.line_items))

        # Validation status with color
        status = inv.validation.status.value if inv.validation else "N/A"
        status_cell = ws.cell(row=row, column=13, value=status)
        if inv.validation:
            status_cell.fill = _status_fill(inv.validation.status)

        ws.cell(row=row, column=14, value=f"{inv.overall_confidence:.0%}")
        ws.cell(row=row, column=15, value="; ".join(inv.processing_errors[:3]))

        # Alternating row colors
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    _auto_width(ws)


# ─── Individual Invoice Sheet ─────────────────────────────────────────────────

def _write_invoice_sheet(ws, inv: Invoice, sheet_name: str):
    """Write detailed invoice data to a sheet in a highly professional dashboard layout."""
    ws.title = sheet_name[:31]

    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True

    # Color Fills
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    section_fill = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid')
    alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    
    # Fonts
    title_font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
    section_font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
    bold_label = Font(name='Calibri', bold=True, size=10, color='333333')
    value_font = Font(name='Calibri', size=10)
    total_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    
    # Borders
    thin_border_side = Side(style='thin', color='D3D3D3')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(style='medium', color='1F4E79'))
    double_bottom = Border(top=Side(style='thin', color='D3D3D3'), bottom=Side(style='double', color='1F4E79'))

    # ─── TITLE BLOCK ───
    ws.merge_cells('A1:L1')
    title_cell = ws.cell(row=1, column=1, value="INVOICE REVIEW DASHBOARD")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = section_fill
    ws.row_dimensions[1].height = 35
    
    # ─── SECTION 1: INVOICE METADATA & EXTRACTION STATS (SIDE-BY-SIDE) ───
    ws.cell(row=3, column=1, value="METADATA").font = section_font
    ws.cell(row=3, column=1).border = thick_bottom
    ws.cell(row=3, column=7, value="EXTRACTION METRICS").font = section_font
    ws.cell(row=3, column=7).border = thick_bottom
    
    metadata_left = [
        ("Invoice Number", str(inv.invoice_number) if str(inv.invoice_number) != "None" else "—"),
        ("Invoice Date", str(inv.invoice_date) if str(inv.invoice_date) != "None" else "—"),
        ("Due Date", inv.due_date or "—"),
        ("PO Number", inv.po_number or "—"),
        ("Place of Supply", inv.place_of_supply or "—"),
        ("Invoice Type", inv.invoice_type or "Tax Invoice"),
        ("Reverse Charge", inv.reverse_charge or "—"),
    ]
    
    metadata_right = [
        ("PDF Type", inv.pdf_type.value.upper()),
        ("Extraction Method", inv.extraction_method.value.upper()),
        ("Overall Confidence", f"{inv.overall_confidence:.1%}"),
        ("Validation Status", inv.validation.status.value if inv.validation else "N/A"),
        ("IRN", inv.irn or "—"),
        ("Ack Number", inv.ack_number or "—"),
        ("Ack Date", inv.ack_date or "—"),
    ]
    
    # Write metadata side-by-side
    for idx, (label, val) in enumerate(metadata_left):
        r = 4 + idx
        ws.cell(row=r, column=1, value=label).font = bold_label
        ws.cell(row=r, column=2, value=val).font = value_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        
    for idx, (label, val) in enumerate(metadata_right):
        r = 4 + idx
        ws.cell(row=r, column=7, value=label).font = bold_label
        ws.cell(row=r, column=8, value=val).font = value_font
        ws.cell(row=r, column=7).border = thin_border
        ws.cell(row=r, column=8).border = thin_border

    # ─── SECTION 2: SELLER & BUYER DETAILS (SIDE-BY-SIDE) ───
    start_r = 12
    ws.cell(row=start_r, column=1, value="SELLER / VENDOR").font = section_font
    ws.cell(row=start_r, column=1).border = thick_bottom
    ws.cell(row=start_r, column=7, value="BUYER / CUSTOMER").font = section_font
    ws.cell(row=start_r, column=7).border = thick_bottom
    
    party_rows = [
        ("Name", inv.seller.name, inv.buyer.name),
        ("GSTIN", inv.seller.gstin, inv.buyer.gstin),
        ("Address", inv.seller.address, inv.buyer.address),
        ("State / Code", f"{inv.seller.state} (Code: {inv.seller.state_code})", f"{inv.buyer.state} (Code: {inv.buyer.state_code})"),
        ("PAN", inv.seller.pan, inv.buyer.pan),
        ("Email / Phone", f"{inv.seller.email} / {inv.seller.phone}" if (inv.seller.email or inv.seller.phone) else "—", f"{inv.buyer.email} / {inv.buyer.phone}" if (inv.buyer.email or inv.buyer.phone) else "—"),
    ]
    
    for idx, (label, s_val, b_val) in enumerate(party_rows):
        r = start_r + 1 + idx
        # Seller
        ws.cell(row=r, column=1, value=label).font = bold_label
        ws.cell(row=r, column=2, value=s_val or "—").font = value_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        # Buyer
        ws.cell(row=r, column=7, value=label).font = bold_label
        ws.cell(row=r, column=8, value=b_val or "—").font = value_font
        ws.cell(row=r, column=7).border = thin_border
        ws.cell(row=r, column=8).border = thin_border

    # ─── SECTION 3: LINE ITEMS TABLE ───
    start_r = 20
    ws.cell(row=start_r, column=1, value="LINE ITEMS").font = section_font
    ws.cell(row=start_r, column=1).border = thick_bottom
    start_r += 1
    
    item_headers = ["Sr", "Description", "HSN/SAC", "Qty", "Unit", "Rate (₹)",
                    "Taxable Amt (₹)", "GST %", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)"]
    for col, h in enumerate(item_headers, 1):
        cell = ws.cell(row=start_r, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[start_r].height = 25
    r = start_r + 1
    
    for i, item in enumerate(inv.line_items):
        ws.cell(row=r, column=1, value=item.sr_no or i + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=item.description)
        ws.cell(row=r, column=3, value=item.hsn_sac or "—").alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=item.quantity).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=item.unit or "—").alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=6, value=item.unit_price).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=7, value=item.taxable_amount).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=8, value=f"{item.gst_rate}%").alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=9, value=item.cgst_amount).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=10, value=item.sgst_amount).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=11, value=item.igst_amount).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=12, value=item.total_amount).number_format = CURRENCY_FORMAT
        
        for col in range(1, len(item_headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = value_font
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = alt_fill
        r += 1

    # Totals Section
    r += 1
    ws.cell(row=r, column=6, value="Subtotal:").font = bold_label
    ws.cell(row=r, column=7, value=inv.subtotal).number_format = CURRENCY_FORMAT
    ws.cell(row=r, column=7).font = value_font
    r += 1
    ws.cell(row=r, column=6, value="Total Tax:").font = bold_label
    ws.cell(row=r, column=7, value=inv.total_tax).number_format = CURRENCY_FORMAT
    ws.cell(row=r, column=7).font = value_font
    r += 1
    ws.cell(row=r, column=6, value="Round Off:").font = bold_label
    ws.cell(row=r, column=7, value=inv.round_off).number_format = CURRENCY_FORMAT
    ws.cell(row=r, column=7).font = value_font
    r += 1
    ws.cell(row=r, column=6, value="Grand Total:").font = total_font
    ws.cell(row=r, column=6).border = double_bottom
    total_val_cell = ws.cell(row=r, column=7, value=inv.grand_total_float)
    total_val_cell.font = total_font
    total_val_cell.number_format = CURRENCY_FORMAT
    total_val_cell.border = double_bottom
    
    r += 2
    if inv.amount_in_words:
        ws.cell(row=r, column=1, value=f"Amount in words: {inv.amount_in_words}").font = bold_label
        r += 2

    # ─── SECTION 4: TAX BREAKDOWN & BANK DETAILS (SIDE-BY-SIDE) ───
    tax_bank_start_r = r
    
    # Headers
    ws.cell(row=tax_bank_start_r, column=1, value="TAX BREAKDOWN").font = section_font
    ws.cell(row=tax_bank_start_r, column=1).border = thick_bottom
    
    ws.cell(row=tax_bank_start_r, column=7, value="BANK DETAILS").font = section_font
    ws.cell(row=tax_bank_start_r, column=7).border = thick_bottom
    
    tax_bank_start_r += 1
    tax_row = tax_bank_start_r
    
    # Write Tax Table
    if inv.tax_details:
        tax_headers = ["Tax Type", "Rate (%)", "Taxable Amount", "Tax Amount"]
        for col_idx, h in enumerate(tax_headers, 1):
            cell = ws.cell(row=tax_bank_start_r, column=col_idx, value=h)
            cell.font = Font(name='Calibri', bold=True, size=10, color='1F4E79')
            cell.fill = section_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        tax_row = tax_bank_start_r + 1
        for td in inv.tax_details:
            ws.cell(row=tax_row, column=1, value=td.tax_type).font = value_font
            ws.cell(row=tax_row, column=2, value=f"{td.rate}%").font = value_font
            ws.cell(row=tax_row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=tax_row, column=3, value=td.taxable_amount).number_format = CURRENCY_FORMAT
            ws.cell(row=tax_row, column=3).font = value_font
            ws.cell(row=tax_row, column=4, value=td.tax_amount).number_format = CURRENCY_FORMAT
            ws.cell(row=tax_row, column=4).font = value_font
            for c in range(1, 5):
                ws.cell(row=tax_row, column=c).border = thin_border
            tax_row += 1

    # Write Bank Details
    bank_row = tax_bank_start_r
    if inv.bank.bank_name:
        bank_rows = [
            ("Bank Name", inv.bank.bank_name),
            ("Account No", inv.bank.account_number),
            ("IFSC Code", inv.bank.ifsc_code),
            ("Branch", inv.bank.branch),
            ("SWIFT Code", inv.bank.swift_code),
        ]
        for label, val in bank_rows:
            if val:
                ws.cell(row=bank_row, column=7, value=label).font = bold_label
                ws.cell(row=bank_row, column=8, value=val).font = value_font
                ws.cell(row=bank_row, column=7).border = thin_border
                ws.cell(row=bank_row, column=8).border = thin_border
                bank_row += 1

    # ─── SECTION 5: VALIDATION ISSUES ───
    if inv.validation and inv.validation.issues:
        r = max(tax_row if inv.tax_details else tax_bank_start_r, bank_row if inv.bank.bank_name else tax_bank_start_r) + 2
        ws.cell(row=r, column=1, value="VALIDATION WARNINGS & LOGS").font = section_font
        ws.cell(row=r, column=1).border = thick_bottom
        r += 1
        
        for issue in inv.validation.issues:
            severity_cell = ws.cell(row=r, column=1, value=issue.severity.value)
            severity_cell.font = bold_label
            severity_cell.fill = _status_fill(issue.severity)
            severity_cell.alignment = Alignment(horizontal='center')
            severity_cell.border = thin_border
            
            field_cell = ws.cell(row=r, column=2, value=issue.field_name)
            field_cell.font = bold_label
            field_cell.border = thin_border
            
            issue_cell = ws.cell(row=r, column=3, value=issue.issue)
            issue_cell.font = value_font
            issue_cell.border = thin_border
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)
            
            r += 1
            
    _auto_width(ws)

# ─── Journal Entry Format (Accuron-specific) ──────────────────────────────────

JOURNAL_HEADERS = [
    "L id", "Acc Period", "Trans Date", "Account Code", "Description",
    "Curr Code", "Trans Amount", "Dr_Cr", "Jrnal Type", "Jrnal Source",
    "Reference", "Description", "Asset Code", "Asset Indicator",
    "Asset / Item Qty", "Due Date", "Branch Analysis Code",
    "Product  Analysis Code", "ChannelAnalysisCode",
    "Sub-Channel  Analysis Code", "Underwriting Year  Analysis Code",
    "Employee Code  Analysis Code", "TDS Applicability  Analysis Code",
    "Department   Analysis Code", "Sequence Code  Analysis Code",
    "Vendor Code  Analysis Code", "Invoice Date", "From Date", "To Date",
    "Addl Date 4", "Addl Date 5", "Cheque & NEFT Number",
    "Invoice Number", "Additional Remarks", "Àdditional Remarks 2",
    "CREDENCE DESCRIPTION", "HSN/SAC NO", "Taxable on Amount",
    "Reverse Charge (Y/N)", "Reverse charge %", "Item Details (Sr.No)",
    "Goods/Service", "GST Tax Rate",
    "Orginal Invoice no for Dr/Cr Notes", "Advance Challan No ",
    "Addl Description 15", "Addl Description 16", "Addl Description 17",
    "Addl Description 18", "Addl Description 19", "Addl Description 20",
    "Addl Description 21", "Addl Description 22", "Addl Description 23",
    "Addl Description 24", "Addl Description 25"
]


def _write_journal_entry_sheet(ws, invoices: list[Invoice]):
    """
    Generate journal entries in Accuron's ERP upload format.
    
    Each invoice generates multiple rows matching Sheet 2 of the template:
    1. Expense line item entries (Debit) - grouped by GST rate and HSN/SAC
    2. SGST input credit entries (Debit) - if applicable, grouped
    3. CGST input credit entries (Debit) - if applicable, grouped
    4. IGST input credit entries (Debit) - if applicable, grouped
    5. Vendor payable base credit row (Credit - GST02)
    6. Vendor payable tax credit row (Credit - GST01)
    """
    ws.title = "upload format"
    ws.views.sheetView[0].showGridLines = True

    # Write headers
    for col, h in enumerate(JOURNAL_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_erp_header_row(ws, 1, len(JOURNAL_HEADERS))
    ws.freeze_panes = 'A2'


    row = 2

    for inv in invoices:
        inv_num = str(inv.invoice_number) if str(inv.invoice_number) != "None" else ""
        inv_date = str(inv.invoice_date) if str(inv.invoice_date) != "None" else ""
        seller_gstin = inv.seller.gstin or ""
        buyer_gstin = inv.buyer.gstin or ""
        seller_name = inv.seller.name or ""

        # Determine accounting period and trans date
        acc_period = ""
        trans_date_val = None
        try:
            from dateutil import parser as dp
            parsed_date = dp.parse(inv_date, dayfirst=True, fuzzy=True)
            trans_date_val = parsed_date
            month = parsed_date.month
            year = parsed_date.year
            # Indian financial year: Apr=001, Mar=012
            fy_start = year if month >= 4 else year - 1
            period_num = month - 3 if month >= 4 else month + 9
            acc_period = f"{fy_start}/{period_num:03d}"
        except Exception:
            trans_date_val = inv_date

        # Resolve vendor code dynamically from CSV
        vendor_code = "CMS0019487"  # Default fallback
        gl_code = 1310020100        # Default fallback
        journal_type = "SEXPS"      # Default fallback
        
        s_name_upper = seller_name.upper()
        
        vendors = _load_vendor_master()
        matched = False
        for v in vendors:
            if v.get('Keyword', '').upper() in s_name_upper:
                vendor_code = v.get('Vendor_Code', vendor_code)
                gl_code = int(v.get('GL_Code', gl_code))
                journal_type = v.get('Journal_Type', journal_type)
                matched = True
                break
                
        if not matched and seller_gstin:
            import hashlib
            stable_hash = int(hashlib.md5(seller_gstin.encode()).hexdigest(), 16) % 10000000
            vendor_code = f"CMS{stable_hash:07d}"
        elif not matched:
            import hashlib
            stable_hash = int(hashlib.md5(seller_name.encode()).hexdigest(), 16) % 10000000
            vendor_code = f"CMS{stable_hash:07d}"
        # Resolve Col L dynamic description e.g., 'APR.26/HK MATERIAL/PORUR'
        desc_period = "APR.26"
        try:
            if trans_date_val:
                desc_period = trans_date_val.strftime("%b.%y").upper()
        except Exception:
            pass

        desc_details = "EXPENSES"
        if inv.line_items:
            first_item_desc = inv.line_items[0].description.upper()
            clean_item_desc = "".join(c if c.isalnum() or c in (" ", "/") else "" for c in first_item_desc)
            desc_details = "/".join(w for w in clean_item_desc.split() if w)[:20]
        col_l_desc = f"{desc_period}/{desc_details}"

        # Determine tax type (IGST if inter-state, CGST+SGST if intra-state)
        is_igst = any(td.tax_type == "IGST" for td in inv.tax_details)

        # If no tax details but place of supply is different, infer IGST
        if not is_igst and inv.place_of_supply and inv.seller.state:
            is_igst = inv.place_of_supply.strip().lower() != inv.seller.state.strip().lower()

        # Group line items by GST rate and HSN/SAC
        # Key: (gst_rate, hsn_sac)
        slab_groups: dict[tuple[float, str], list] = {}
        for item in inv.line_items:
            rate = item.gst_rate
            hsn = str(item.hsn_sac or "").strip()
            key = (rate, hsn)
            if key not in slab_groups:
                slab_groups[key] = []
            slab_groups[key].append(item)

        # Fallback: if no line items but we parsed tax_details, create synthetic group
        if not slab_groups and inv.tax_details:
            for td in inv.tax_details:
                rate = td.rate if "IGST" in td.tax_type else td.rate * 2
                slab_groups[(rate, "")] = []

        total_base = 0.0
        total_tax = 0.0
        is_first_invoice_row = True

        # Track the rows to write
        # Each group generates Debit rows
        for (gst_rate, hsn_str), items in slab_groups.items():
            group_taxable = sum(i.taxable_amount for i in items) if items else inv.subtotal
            if group_taxable <= 0 and len(slab_groups) == 1:
                group_taxable = inv.subtotal
            
            group_taxable_rounded = round(group_taxable, 2)
            group_desc = items[0].description if items else "OFFICE EXPENSES"
            total_base += group_taxable_rounded

            # ─── Row 1: Expense entry (Debit) ───
            # Col 1: L id (first row of invoice gets '1;3;6;7', others '6;7')
            l_id = "1;3;6;7" if is_first_invoice_row else "6;7"
            is_first_invoice_row = False

            # Column mappings: 0-indexed internally
            ws.cell(row=row, column=1, value=l_id)
            ws.cell(row=row, column=2, value=acc_period)
            ws.cell(row=row, column=3, value=trans_date_val)
            ws.cell(row=row, column=4, value=gl_code)  # Expense GL
            ws.cell(row=row, column=5, value="REPAIRS AND MAINT. OFFICE BLDG")
            ws.cell(row=row, column=6, value="INR")
            ws.cell(row=row, column=7, value=group_taxable_rounded)
            ws.cell(row=row, column=8, value="D")
            ws.cell(row=row, column=9, value=journal_type)
            ws.cell(row=row, column=11, value=f"{vendor_code}/{seller_name.upper()}")
            ws.cell(row=row, column=12, value=col_l_desc)
            ws.cell(row=row, column=17, value="LXS00")
            ws.cell(row=row, column=18, value="_NA")
            ws.cell(row=row, column=19, value="CC004")
            ws.cell(row=row, column=20, value="_NA")
            ws.cell(row=row, column=21, value="_NA")
            ws.cell(row=row, column=22, value="_NA")
            ws.cell(row=row, column=23, value="TD02")  # TDS Applicability
            ws.cell(row=row, column=24, value="SRO")
            ws.cell(row=row, column=26, value=vendor_code)
            ws.cell(row=row, column=27, value=trans_date_val)
            ws.cell(row=row, column=32, value=seller_gstin)
            ws.cell(row=row, column=33, value=inv_num)
            
            # Reverse charge (AM, AN) -> columns 39, 40
            rev_charge = inv.reverse_charge if hasattr(inv, 'reverse_charge') else "No"
            ws.cell(row=row, column=39, value=rev_charge)
            ws.cell(row=row, column=40, value=rev_charge)
            
            ws.cell(row=row, column=44, value=buyer_gstin)
            row += 1

            # ─── GST Debit Rows (if rate > 0) ───
            if gst_rate > 0:
                if is_igst:
                    # IGST Row (Debit)
                    igst_amount = round(group_taxable * gst_rate / 100, 2)
                    total_tax += igst_amount

                    ws.cell(row=row, column=1, value="6;7")
                    ws.cell(row=row, column=2, value=acc_period)
                    ws.cell(row=row, column=3, value=trans_date_val)
                    ws.cell(row=row, column=4, value=1120599006)  # IGST INPUT GL
                    ws.cell(row=row, column=5, value="GST input cr exps-IGST")
                    ws.cell(row=row, column=6, value="INR")
                    ws.cell(row=row, column=7, value=igst_amount)
                    ws.cell(row=row, column=8, value="D")
                    ws.cell(row=row, column=9, value=journal_type)
                    ws.cell(row=row, column=11, value=f"{vendor_code}/{seller_name.upper()}")
                    ws.cell(row=row, column=12, value=col_l_desc)
                    ws.cell(row=row, column=17, value="LXS00")
                    ws.cell(row=row, column=18, value="_NA")
                    ws.cell(row=row, column=19, value="CC004")
                    ws.cell(row=row, column=20, value="_NA")
                    ws.cell(row=row, column=21, value="_NA")
                    ws.cell(row=row, column=22, value="_NA")
                    ws.cell(row=row, column=23, value="_NA")
                    ws.cell(row=row, column=24, value="SRO")
                    ws.cell(row=row, column=26, value=vendor_code)
                    ws.cell(row=row, column=27, value=trans_date_val)
                    ws.cell(row=row, column=32, value=seller_gstin)
                    ws.cell(row=row, column=33, value=inv_num)
                    ws.cell(row=row, column=37, value=hsn_str)
                    ws.cell(row=row, column=38, value=group_taxable_rounded)
                    ws.cell(row=row, column=43, value=gst_rate)
                    ws.cell(row=row, column=44, value=buyer_gstin)
                    ws.cell(row=row, column=45, value=1310020100)  # Offset Expense GL
                    row += 1
                else:
                    # SGST + CGST Rows (Debit)
                    tax_amount = round(group_taxable * gst_rate / 200, 2)
                    total_tax += tax_amount * 2

                    # SGST Row
                    ws.cell(row=row, column=1, value="6;7")
                    ws.cell(row=row, column=2, value=acc_period)
                    ws.cell(row=row, column=3, value=trans_date_val)
                    ws.cell(row=row, column=4, value=1120599005)  # SGST INPUT GL
                    ws.cell(row=row, column=5, value="GST input cr exps-SGST")
                    ws.cell(row=row, column=6, value="INR")
                    ws.cell(row=row, column=7, value=tax_amount)
                    ws.cell(row=row, column=8, value="D")
                    ws.cell(row=row, column=9, value=journal_type)
                    ws.cell(row=row, column=11, value=f"{vendor_code}/{seller_name.upper()}")
                    ws.cell(row=row, column=12, value=col_l_desc)
                    ws.cell(row=row, column=17, value="LXS00")
                    ws.cell(row=row, column=18, value="_NA")
                    ws.cell(row=row, column=19, value="CC004")
                    ws.cell(row=row, column=20, value="_NA")
                    ws.cell(row=row, column=21, value="_NA")
                    ws.cell(row=row, column=22, value="_NA")
                    ws.cell(row=row, column=23, value="_NA")
                    ws.cell(row=row, column=24, value="SRO")
                    ws.cell(row=row, column=26, value=vendor_code)
                    ws.cell(row=row, column=27, value=trans_date_val)
                    ws.cell(row=row, column=32, value=seller_gstin)
                    ws.cell(row=row, column=33, value=inv_num)
                    ws.cell(row=row, column=37, value=hsn_str)
                    ws.cell(row=row, column=38, value=group_taxable_rounded)
                    ws.cell(row=row, column=43, value=gst_rate / 2)
                    ws.cell(row=row, column=44, value=buyer_gstin)
                    ws.cell(row=row, column=45, value=1310020100)
                    row += 1

                    # CGST Row
                    ws.cell(row=row, column=1, value="6;7")
                    ws.cell(row=row, column=2, value=acc_period)
                    ws.cell(row=row, column=3, value=trans_date_val)
                    ws.cell(row=row, column=4, value=1120599007)  # CGST INPUT GL
                    ws.cell(row=row, column=5, value="GST input cr exps-CGST")
                    ws.cell(row=row, column=6, value="INR")
                    ws.cell(row=row, column=7, value=tax_amount)
                    ws.cell(row=row, column=8, value="D")
                    ws.cell(row=row, column=9, value=journal_type)
                    ws.cell(row=row, column=11, value=f"{vendor_code}/{seller_name.upper()}")
                    ws.cell(row=row, column=12, value=col_l_desc)
                    ws.cell(row=row, column=17, value="LXS00")
                    ws.cell(row=row, column=18, value="_NA")
                    ws.cell(row=row, column=19, value="CC004")
                    ws.cell(row=row, column=20, value="_NA")
                    ws.cell(row=row, column=21, value="_NA")
                    ws.cell(row=row, column=22, value="_NA")
                    ws.cell(row=row, column=23, value="_NA")
                    ws.cell(row=row, column=24, value="SRO")
                    ws.cell(row=row, column=26, value=vendor_code)
                    ws.cell(row=row, column=27, value=trans_date_val)
                    ws.cell(row=row, column=32, value=seller_gstin)
                    ws.cell(row=row, column=33, value=inv_num)
                    ws.cell(row=row, column=37, value=hsn_str)
                    ws.cell(row=row, column=38, value=group_taxable_rounded)
                    ws.cell(row=row, column=43, value=gst_rate / 2)
                    ws.cell(row=row, column=44, value=buyer_gstin)
                    ws.cell(row=row, column=45, value=1310020100)
                    row += 1

        # ─── Row 4: Vendor base Credit (Credit) ───
        ws.cell(row=row, column=1, value="6;7")
        ws.cell(row=row, column=2, value=acc_period)
        ws.cell(row=row, column=3, value=trans_date_val)
        ws.cell(row=row, column=4, value=vendor_code)
        ws.cell(row=row, column=5, value=seller_name)
        ws.cell(row=row, column=6, value="INR")
        ws.cell(row=row, column=7, value=round(total_base, 2))
        ws.cell(row=row, column=8, value="C")
        ws.cell(row=row, column=9, value="SEXPS")
        ws.cell(row=row, column=11, value=inv.source_file.split('.')[0][:20])
        ws.cell(row=row, column=12, value=col_l_desc)
        ws.cell(row=row, column=17, value="GST02")  # Base credit code
        ws.cell(row=row, column=18, value="_NA")
        ws.cell(row=row, column=19, value="CC004")
        ws.cell(row=row, column=20, value="_NA")
        ws.cell(row=row, column=21, value="_NA")
        ws.cell(row=row, column=22, value="_NA")
        ws.cell(row=row, column=23, value="_NA")
        ws.cell(row=row, column=24, value="SRO")
        ws.cell(row=row, column=26, value=vendor_code)
        ws.cell(row=row, column=27, value=trans_date_val)
        ws.cell(row=row, column=32, value=seller_gstin)
        ws.cell(row=row, column=33, value=inv_num)
        ws.cell(row=row, column=44, value=buyer_gstin)
        row += 1

        # ─── Row 5: Vendor tax Credit (Credit) ───
        if total_tax > 0:
            ws.cell(row=row, column=1, value="6;7")
            ws.cell(row=row, column=2, value=acc_period)
            ws.cell(row=row, column=3, value=trans_date_val)
            ws.cell(row=row, column=4, value=vendor_code)
            ws.cell(row=row, column=5, value=seller_name)
            ws.cell(row=row, column=6, value="INR")
            ws.cell(row=row, column=7, value=round(total_tax, 2))
            ws.cell(row=row, column=8, value="C")
            ws.cell(row=row, column=9, value="SEXPS")
            ws.cell(row=row, column=11, value=inv.source_file.split('.')[0][:20])
            ws.cell(row=row, column=12, value=col_l_desc)
            ws.cell(row=row, column=17, value="GST01")  # Tax credit code
            ws.cell(row=row, column=18, value="_NA")
            ws.cell(row=row, column=19, value="CC004")
            ws.cell(row=row, column=20, value="_NA")
            ws.cell(row=row, column=21, value="_NA")
            ws.cell(row=row, column=22, value="_NA")
            ws.cell(row=row, column=23, value="_NA")
            ws.cell(row=row, column=24, value="SRO")
            ws.cell(row=row, column=26, value=vendor_code)
            ws.cell(row=row, column=27, value=trans_date_val)
            ws.cell(row=row, column=32, value=seller_gstin)
            ws.cell(row=row, column=33, value=inv_num)
            ws.cell(row=row, column=44, value=buyer_gstin)
            row += 1

        # Empty row separator between invoices
        row += 1

    _auto_width(ws)


DETAILS_CAPTURED_HEADERS = ["Head", "Details"]

DETAILS_CAPTURED_ROWS = [
    ("Acc Period", "Period of entry"),
    ("Trans Date", "Date of entry"),
    ("Account Code", "Respective GL code"),
    ("Curr Code", "INR"),
    ("Trans Amount", "Amount"),
    ("Dr_Cr", "Debit or Credit"),
    ("Jrnal Type", "SEXPS or HEXPS"),
    ("Jrnal Source", "Not applicable"),
    ("Reference", "Vendor code and Name"),
    ("Description", "Nature of expenses"),
    ("Asset Code", "to be blank"),
    ("Asset Indicator", "to be blank"),
    ("Asset / Item Qty", "to be blank"),
    ("Due Date", "to be blank"),
    ("Branch Analysis Code", "Branch code"),
    ("Product  Analysis Code", "Not applicable"),
    ("ChannelAnalysisCode", "Respective channel code"),
    ("Sub-Channel  Analysis Code", "Not applicable"),
    ("Underwriting Year  Analysis Code", "Not applicable"),
    ("Employee Code  Analysis Code", "Not applicable"),
    ("TDS Applicability  Analysis Code", "TD01  if TDS applicable"),
    ("Department   Analysis Code", None),
    ("Sequence Code  Analysis Code", "Not applicable"),
    ("Vendor Code  Analysis Code", "Vendor code"),
    ("Invoice Date", "Invoice date"),
    ("From Date", "Not applicable"),
    ("To Date", "Not applicable"),
    ("Addl Date 4", None),
    ("Addl Date 5", None),
    ("Cheque & NEFT Number", "Vendor GST number"),
    ("Invoice Number", "Vendor invoice number"),
    ("Additional Remarks", "NEFT or Cheque"),
    ("Àdditional Remarks 2", "Not applicable"),
    ("CREDENCE DESCRIPTION", "Not applicable"),
    ("HSN/SAC NO", "As per invoice"),
    ("Taxable on Amount", "Invoice value for GST"),
    ("Reverse Charge (Y/N)", "Not applicable"),
    ("Reverse charge %", "Not applicable"),
    ("Item Details (Sr.No)", "Not applicable"),
    ("Goods/Service", "Not applicable"),
    ("GST Tax Rate", "GST rate"),
    ("Orginal Invoice no for Dr/Cr Notes", "RS GST number"),
    ("Advance Challan No ", "Respective expenses account code"),
]


def _write_details_captured_sheet(ws):
    """Write the reference guide sheet matching the template's first sheet."""
    ws.title = "details to be captured"
    ws.views.sheetView[0].showGridLines = True
    
    # Write headers
    for col, h in enumerate(DETAILS_CAPTURED_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        
    ws.row_dimensions[1].height = 25
    
    # Write rows
    for r_idx, (head, details) in enumerate(DETAILS_CAPTURED_ROWS, 2):
        cell_head = ws.cell(row=r_idx, column=1, value=head)
        cell_details = ws.cell(row=r_idx, column=2, value=details)
        
        cell_head.font = Font(name="Calibri", size=11)
        cell_details.font = Font(name="Calibri", size=11)
        
        cell_head.border = THIN_BORDER
        cell_details.border = THIN_BORDER
        
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35


# ─── Main Generator ──────────────────────────────────────────────────────────

def generate_workbook(invoices: list[Invoice], output_path: str,
                      include_journal: bool = True):
    """
    Generate the complete Excel workbook.
    
    Sheets:
    1. Summary — all invoices overview
    2. Individual invoice tabs (named by invoice number)
    3. details to be captured — template column reference sheet (optional)
    4. upload format — Accuron ERP upload format (optional)
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    _write_summary_sheet(ws_summary, invoices)

    # Sheets 2-N: Individual invoices
    for i, inv in enumerate(invoices, 1):
        inv_num = str(inv.invoice_number).replace("/", "-")[:20] or f"Invoice-{i}"
        ws = wb.create_sheet()
        _write_invoice_sheet(ws, inv, inv_num)

    # Sheets N+1 and N+2: Accuron ERP template sheets
    if include_journal:
        ws_details = wb.create_sheet()
        _write_details_captured_sheet(ws_details)

        ws_journal = wb.create_sheet()
        _write_journal_entry_sheet(ws_journal, invoices)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel workbook saved: {output_path}")
    return output_path

